from typing import Literal
from pathlib import Path
import pandas as pd

from dotools_py.logger import logger
from dotools_py._utils import convert_path
from dotools_py._custom_class import PathLike, EmptyType, InputError
from dotools_py.io._utils import _check_backend

_Empty = EmptyType()

def _check_sheet_name(path: PathLike):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, keep_links=False)
    return wb.sheetnames

def read_excel(
    path: PathLike,
    filename: str | None = None,
    sheet_name: str = "Sheet 1",
    backend: Literal["pandas", "polars"] = "pandas",
    **kwargs
) -> pd.DataFrame:
    """Read Excel Sheet into a DataFrame.

    Parameters
    ----------
    path
        Directory containing the Excel Sheet.
    filename
        Name of the Excel Sheet file, including its extension. If not specified, assume that `path` contains the full path to the ExcelSheet.
    sheet_name
        Name of the Sheet to read.
    backend
        Library to use for reading. If ``"polars"`` is selected and reading fails, Pandas is used as a fallback.
    **kwargs
        Additional arguments passed directly to
        `polars.read_excel <https://docs.pola.rs/api/python/stable/reference/api/polars.read_excel.html>`_
        or
        `pandas.read_excel <https://pandas.pydata.org/docs/reference/api/pandas.read_excel.html>`_.

    Returns
    -------
    Returns a `pd.DataFrame` containing the content from the selected sheet.

    """
    import polars as pl

    _check_backend(backend, ["pandas","polars"])
    input_path = convert_path(path) if filename is None else convert_path(path) / filename
    df = _Empty

    sheets = _check_sheet_name(input_path)
    if len(sheets) == 1:
        sheets = sheets[0]
        if sheets != sheet_name:
            logger.warn(f"{sheet_name} is not present, using {sheets}")
            sheet_name = sheets
    else:
        if sheet_name not in sheets:
            raise InputError(f"{sheet_name} is not a valid Sheet. Available sheets: {sheets}")

    if backend == "polars":
        try:
            df = pl.read_excel(source=input_path, sheet_name=sheet_name, **kwargs)
            df = df.to_pandas()
        except Exception as e:
            logger.warn(f"Error using polars backend falling back to pandas.\n{e}")
    if df is _Empty:
        df = pd.read_excel(io=input_path, sheet_name=sheet_name, **kwargs)

    if "Unnamed: 0" in df.columns:
        df.set_index("Unnamed: 0", inplace=True)
        df.index.name = None
    if "" in df.columns:
        del df[""]
    return df


def read_csv(
    path: PathLike,
    filename: str | None = None,
    delimiter: str = ",",
    backend: Literal["pandas", "polars"] = "pandas",
    **kwargs
) -> pd.DataFrame:
    """Read comma separated files into a DataFrame.

    Parameters
    ----------
    path
        Directory containing the comma separated file.
    filename
        Name of the comma separated file. If not specified, assume that `path` contains the full path to the file.
    delimiter
        Character or regex pattern to treat as the delimiter.
    backend
        Library to use for reading. If ``"polars"`` is selected and reading fails, Pandas is used as a fallback.
    **kwargs
        Additional arguments passed directly to
            `polars.read_csv <https://docs.pola.rs/api/python/stable/reference/api/polars.read_csv.html>`_
            or
            `pandas.read_csv <https://pandas.pydata.org/docs/reference/api/pandas.read_csv.html>`_.

    Returns
    -------
    Returns a `pd.DataFrame` containing the content from the selected sheet.

    """
    import polars as pl

    _check_backend(backend, ["pandas","polars"])
    input_path = convert_path(path) if filename is None else convert_path(path) / filename
    df  = _Empty

    if backend == "polars":
        try:
            df = pl.read_csv(source=input_path, separator=delimiter, **kwargs)
            df = df.to_pandas()
        except Exception as e:
            logger.warn(f"Error using polars backend falling back to pandas.\n{e}")

    if df is _Empty:
        df = pd.read_csv(input_path, sep=delimiter, iterator=False, **kwargs)
    if "" in df.columns:
        del df[""]
    if "Unnamed: 0" in df.columns:
        df.set_index("Unnamed: 0", inplace=True)
        df.index.name = None
    return  df


def read_parquet(
    path: PathLike,
    filename: str | None = None,
    backend: Literal["pandas", "polars"] = "pandas",
    **kwargs
) -> pd.DataFrame:
    """Read a parquet object into a DataFrame.

    Parameters
    ----------
    path
        Directory containing the comma separated file.
    filename
         Name of the parquet file. If not specified, assume that `path` contains the full path to the file.
    backend
        Library to use for reading. If ``"polars"`` is selected and reading fails, Pandas is used as a fallback.
    **kwargs
        Additional arguments passed directly to
            `polars.read_parquet <https://docs.pola.rs/api/python/stable/reference/api/polars.read_parquet.html>`_
            or
            `pandas.read_parquet <https://pandas.pydata.org/docs/reference/api/pandas.read_parquet.html>`_.

    Returns
    -------
    Returns a `pd.DataFrame` containing the content from the selected sheet.

    """
    import polars as pl

    _check_backend(backend, ["pandas","polars"])
    input_path: Path = convert_path(path) if filename is None else convert_path(path) / filename

    df = _Empty

    if backend == "polars":
        try:
            df = pl.read_parquet(source=input_path, **kwargs)
            df = df.to_pandas()
        except Exception as e:
            logger.warn(f"Error using polars backend falling back to pandas.\n{e}")

    if df is _Empty:
        df = pd.read_parquet(input_path, **kwargs)

    if "__index_level_0__" in df.columns:
        df.set_index("__index_level_0__", inplace=True)
        df.index.name = None
    return df

